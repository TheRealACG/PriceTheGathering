"""
Andrew Greer
Price The Gathering

3/30/2019
"""

import os
import csv
import openpyxl

from bs4 import BeautifulSoup
import requests

#global variables, rows and columns 0 indexed
CARD_NAME_COLUMN = 2;
CARD_EXPANSION_COLUMN = 3;
CARD_PRICE_COLUMN = 5;

# path of excel photo
loc = (r"C:\Users\Andrew\Documents\MTG card collection.xlsx")

def cardNameCheck(name):
    name = name.replace(" ", "+")
    name = name.replace(",", "")
    return name
#end of cardNameCheck

def goldfishSearch(url):
    source = requests.get(url).text
    soup = BeautifulSoup(source, 'lxml')
    prices = soup.find("div", {"class": "price-box paper"})
    paperPrice = prices.find("div", {"class": "price-box-price"})
    return paperPrice.text
#end of goldfishSearch


#start of code
wb = openpyxl.load_workbook(loc)
sheet = wb['Sheet1']
print(sheet.max_row)

for record in range(sheet.max_row - 1):
    cardName = sheet.cell(row=(record+2),column=CARD_NAME_COLUMN).value
    print(cardName)

    if cardName == None:
        continue
    cardName = cardNameCheck(cardName)
    expansion = sheet.cell(row=(record+2),column=CARD_EXPANSION_COLUMN).value
    expansion = cardNameCheck(expansion)
    #print(cardName)
    #print(expansion)
    startingURL = 'https://www.mtggoldfish.com/price/' + expansion + '/' + cardName + '#paper'
    cardValue = goldfishSearch(startingURL)
    print(cardValue)

    #insert value into excel
    sheet.cell(row=(record+2),column=CARD_PRICE_COLUMN).value = cardValue

#excel must be closed to run!!!
wb.save(loc)
print("We done!")